"""
Modelo de abasto y transferencias para Google Colab.

Flujo:
1. Autentica al usuario de Colab contra Google Drive.
2. Exporta DATA_TRANSFERS (Google Sheet) como XLSX.
3. Busca en TR_PLANS el CSV DD-MM-YYYY.csv de la fecha de ejecución.
4. Calcula demanda, prioridades, stock ajustado, capacidad y tareas.
5. Genera un reporte XLSX y un CSV independiente por WAREHOUSE_SOURCE.
6. Guarda los resultados en TR_PLANS/SALIDAS/DD-MM-YYYY.

El motor es deliberadamente secuencial en la fase de asignación: el stock,
la capacidad y el número de tareas cambian después de cada requerimiento.
"""

# %% [markdown]
# CONFIGURACIÓN: modifica únicamente este bloque en Google Colab.

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator
from zoneinfo import ZoneInfo
import csv
import io
import math
import re
import shutil
import statistics
import sys
import unicodedata


@dataclass
class Config:
    # Aceptan el ID o la URL completa.
    data_transfers_spreadsheet: str = "18kHevkMvf9l4s6ANg3h5KdNyj2yEPGAp5C_t8JwxFVw"
    tr_plans_folder: str = "1yDJTSClDR1ikSS6-sZNAbCWwtjTf384z"

    # El orden representa la prioridad de consumo del stock.
    origin_warehouses: tuple[int, ...] = (811, 834)
    max_tasks: int = 14_000

    # None usa la fecha actual de Ciudad de México.
    # Para reprocesar: "28-08-2026".
    run_date_override: str | None = None
    timezone: str = "America/Mexico_City"

    default_store_capacity_m3: float = 10.0
    default_m3_per_unit: float = 0.002
    minimum_positive_quantity: int = 3

    # Carpeta remota: TR_PLANS/SALIDAS/DD-MM-YYYY.
    remote_output_root_name: str = "SALIDAS"
    local_work_dir: str = "/content/modelo_abasto"

    # En una segunda ejecución del mismo día actualiza los archivos del día.
    replace_same_day_outputs: bool = True
    generate_empty_source_files: bool = False


CONFIG = Config()


# %% Dependencias

def ensure_dependencies() -> None:
    """Instala únicamente las dependencias que falten en el runtime de Colab."""
    missing: list[str] = []
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl>=3.1")
    try:
        import xlsxwriter  # noqa: F401
    except ImportError:
        missing.append("xlsxwriter>=3.2")
    if missing:
        import subprocess

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q", *missing]
        )


ensure_dependencies()

import openpyxl
import xlsxwriter


OUTPUT_COLUMNS = [
    "WAREHOUSE_DESTINATION",
    "WAREHOUSE_SOURCE",
    "RETAIL_ID",
    "QUANTITY",
    "PLANNED_DATE",
    "ROUTE",
    "DELIVERY_PRIORITY",
    "CITY",
    "STORAGE",
    "VALUE",
]

CDMX = "CDMX"
GDL = "GDL"
MTY = "MTY"
FOREIGN_DESTINATION_CITIES = {GDL, MTY}


# %% Utilidades generales

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def normalize_city(value: Any) -> str:
    raw = clean_text(value)
    folded = "".join(
        char
        for char in unicodedata.normalize("NFKD", raw)
        if not unicodedata.combining(char)
    ).upper()
    folded = re.sub(r"\s+", " ", folded).strip()
    if folded in {"CDMX", "CIUDAD DE MEXICO", "MEXICO CITY"}:
        return CDMX
    if "GUADALAJARA" in folded or folded == "GDL":
        return GDL
    if "MONTERREY" in folded or folded == "MTY":
        return MTY
    return folded


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None or clean_text(value) == "":
        return default
    if isinstance(value, (int, float)):
        result = float(value)
    else:
        text = clean_text(value).replace(",", "")
        try:
            result = float(text)
        except ValueError:
            return default
    return result if math.isfinite(result) else default


def to_id(value: Any, field_name: str, allow_none: bool = False) -> int | None:
    if value is None or clean_text(value) == "":
        if allow_none:
            return None
        raise ValueError(f"{field_name}: identificador vacío")
    number = to_float(value, default=math.nan)
    if not math.isfinite(number) or not math.isclose(number, round(number), abs_tol=1e-6):
        raise ValueError(f"{field_name}: identificador inválido {value!r}")
    return int(round(number))


def parse_run_date(config: Config) -> date:
    if not config.run_date_override:
        return datetime.now(ZoneInfo(config.timezone)).date()
    value = config.run_date_override.strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    raise ValueError(
        "run_date_override debe usar DD-MM-YYYY, YYYY-MM-DD o DD/MM/YYYY"
    )


def extract_drive_id(value: str, label: str) -> str:
    value = clean_text(value)
    if not value or "PEGA_AQUI" in value:
        raise ValueError(f"Falta configurar {label}")
    patterns = (
        r"/spreadsheets/d/([A-Za-z0-9_-]+)",
        r"/folders/([A-Za-z0-9_-]+)",
        r"^([A-Za-z0-9_-]{15,})$",
    )
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            return match.group(1)
    raise ValueError(f"No pude extraer el ID de {label}: {value!r}")


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._")


# %% Google Drive

def authenticate_drive():
    try:
        from google.colab import auth
    except ImportError as exc:
        raise RuntimeError(
            "La autenticación automática está diseñada para Google Colab. "
            "Para pruebas locales usa run_pipeline(..., upload_to_drive=False)."
        ) from exc

    auth.authenticate_user()
    try:
        import googleapiclient  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q", "google-api-python-client>=2.0"]
        )
    import google.auth
    import httplib2
    from google_auth_httplib2 import AuthorizedHttp
    from googleapiclient.discovery import build

    credentials, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    authorized_http = AuthorizedHttp(credentials, http=httplib2.Http(timeout=180))
    return build("drive", "v3", http=authorized_http, cache_discovery=False)


def escape_drive_query(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def list_exact_files(drive, folder_id: str, names: Iterable[str]) -> list[dict[str, Any]]:
    escaped_names = [f"name = '{escape_drive_query(name)}'" for name in names]
    query = (
        f"'{escape_drive_query(folder_id)}' in parents and trashed = false and "
        f"({' or '.join(escaped_names)})"
    )
    response = (
        drive.files()
        .list(
            q=query,
            fields="files(id,name,mimeType,modifiedTime,webViewLink,parents)",
            orderBy="modifiedTime desc",
            pageSize=100,
        )
        .execute()
    )
    return response.get("files", [])


def download_drive_file(drive, file_id: str, destination: Path) -> None:
    from googleapiclient.http import MediaIoBaseDownload

    metadata = drive.files().get(fileId=file_id, fields="id,name,mimeType").execute()
    if metadata["mimeType"] == "application/vnd.google-apps.spreadsheet":
        request = drive.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive.files().get_media(fileId=file_id)

    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("wb") as handle:
        downloader = MediaIoBaseDownload(handle, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()


def find_daily_plan_file(drive, folder_id: str, run_date: date) -> dict[str, Any]:
    stem = run_date.strftime("%d-%m-%Y")
    files = list_exact_files(drive, folder_id, [f"{stem}.csv", f"{stem}.CSV"])
    if not files:
        raise FileNotFoundError(
            f"No existe {stem}.csv ni {stem}.CSV dentro de la carpeta TR_PLANS"
        )
    if len(files) > 1:
        names = ", ".join(f"{f['name']} ({f['id']})" for f in files)
        raise RuntimeError(
            "Encontré más de un plan para la fecha. Elimina la ambigüedad: " + names
        )
    return files[0]


def find_or_create_folder(drive, parent_id: str, name: str) -> dict[str, Any]:
    files = list_exact_files(drive, parent_id, [name])
    folders = [
        item
        for item in files
        if item.get("mimeType") == "application/vnd.google-apps.folder"
    ]
    if len(folders) > 1:
        raise RuntimeError(f"Existen varias carpetas llamadas {name!r} bajo el mismo padre")
    if folders:
        return folders[0]
    return (
        drive.files()
        .create(
            body={
                "name": name,
                "mimeType": "application/vnd.google-apps.folder",
                "parents": [parent_id],
            },
            fields="id,name,mimeType,webViewLink",
        )
        .execute()
    )


def upload_or_replace_file(
    drive,
    local_path: Path,
    folder_id: str,
    mime_type: str,
    replace: bool,
) -> dict[str, Any]:
    from googleapiclient.http import MediaFileUpload

    existing = list_exact_files(drive, folder_id, [local_path.name])
    media = MediaFileUpload(str(local_path), mimetype=mime_type, resumable=True)
    if replace and len(existing) == 1:
        return (
            drive.files()
            .update(
                fileId=existing[0]["id"],
                media_body=media,
                fields="id,name,mimeType,webViewLink,modifiedTime",
            )
            .execute()
        )
    if len(existing) > 1:
        raise RuntimeError(
            f"Hay varios archivos llamados {local_path.name!r} en la carpeta de salida"
        )
    if existing and not replace:
        timestamp = datetime.now().strftime("%H%M%S")
        upload_name = f"{local_path.stem}_{timestamp}{local_path.suffix}"
    else:
        upload_name = local_path.name
    return (
        drive.files()
        .create(
            body={"name": upload_name, "parents": [folder_id]},
            media_body=media,
            fields="id,name,mimeType,webViewLink,modifiedTime",
        )
        .execute()
    )


# %% Lectura de DATA_TRANSFERS

def find_header_row(ws, required_headers: Iterable[str], scan_rows: int = 40) -> tuple[int, dict[str, int]]:
    required = {normalize_header(header) for header in required_headers}
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
    ):
        positions: dict[str, int] = {}
        for column_index, value in enumerate(row):
            header = normalize_header(value)
            if header:
                positions[header] = column_index
        if required.issubset(positions):
            return row_number, positions
    raise ValueError(
        f"No encontré encabezados {sorted(required)} en la pestaña {ws.title!r}"
    )


def iter_sheet_records(
    workbook,
    sheet_name: str,
    required_headers: Iterable[str],
    selected_headers: Iterable[str] | None = None,
) -> Iterator[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Falta la pestaña obligatoria {sheet_name!r}")
    ws = workbook[sheet_name]
    header_row, positions = find_header_row(ws, required_headers)
    selected = list(selected_headers or required_headers)
    normalized_selected = [(header, normalize_header(header)) for header in selected]
    missing = [header for header, norm in normalized_selected if norm not in positions]
    if missing:
        raise ValueError(f"{sheet_name}: faltan columnas {missing}")

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = {
            header: row[positions[norm]] if positions[norm] < len(row) else None
            for header, norm in normalized_selected
        }
        if any(clean_text(value) for value in record.values()):
            yield record


def put_unique(
    target: dict[Any, Any],
    key: Any,
    value: Any,
    warnings: list[str],
    label: str,
    conflict_policy: str = "error",
) -> None:
    if key not in target:
        target[key] = value
        return
    previous = target[key]
    if previous == value:
        return
    message = f"{label}: llave {key!r} tiene valores distintos {previous!r} y {value!r}"
    if conflict_policy == "min":
        target[key] = min(previous, value)
        warnings.append(message + f"; se usó el menor: {target[key]!r}")
    elif conflict_policy == "first":
        warnings.append(message + f"; se conservó el primero: {previous!r}")
    else:
        raise ValueError(message)


def copernico_is_usable(location: Any) -> bool:
    """Replica las fórmulas históricas de ubicación y USABLE?."""
    value = clean_text(location).upper()
    if value.startswith("Z"):
        return True
    if value in {"CANCELADOS", "RECIBO_444"}:
        return False
    # REGEXEXTRACT("(.)(.)(.)(..)(.)(..)") requiere al menos 8 caracteres.
    return len(value) >= 8


def load_copernico_unusable_csv(
    path: Path,
) -> tuple[dict[tuple[int, int], float], dict[str, Any]]:
    """Acumula saldo no usable por la combinación Bodega + producto."""
    header_aliases = {
        "Bodega": {"BODEGA", "WAREHOUSE ID", "WAREHOUSE_ID"},
        "EAN": {"EAN", "PRODUCT ID", "PRODUCT_ID"},
        "Ubicacion": {"UBICACION", "UBICACIÓN", "LOCATION"},
        "Saldo": {"SALDO", "STOCK", "QUANTITY"},
    }

    def normalized_csv_header(value: Any) -> str:
        folded = "".join(
            character
            for character in unicodedata.normalize("NFKD", clean_text(value))
            if not unicodedata.combining(character)
        )
        return re.sub(r"\s+", " ", folded.replace("_", " ")).strip().upper()

    normalized_aliases = {
        canonical: {normalized_csv_header(alias) for alias in aliases}
        for canonical, aliases in header_aliases.items()
    }
    unusable_stock: dict[tuple[int, int], float] = defaultdict(float)
    total_rows = 0
    source_rows = 0
    unusable_rows = 0

    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        if reader.fieldnames is None:
            raise ValueError("El CSV de COPÉRNICO no contiene encabezados")

        normalized_fields = {
            normalized_csv_header(field_name): field_name
            for field_name in reader.fieldnames
            if clean_text(field_name)
        }
        field_lookup: dict[str, str] = {}
        missing: list[str] = []
        for canonical, aliases in normalized_aliases.items():
            match = next(
                (
                    normalized_fields[alias]
                    for alias in aliases
                    if alias in normalized_fields
                ),
                None,
            )
            if match is None:
                missing.append(canonical)
            else:
                field_lookup[canonical] = match
        if missing:
            raise ValueError(
                "El CSV de COPÉRNICO no contiene las columnas obligatorias: "
                + ", ".join(missing)
            )

        for row_number, row in enumerate(reader, start=2):
            total_rows += 1
            warehouse = to_id(
                row.get(field_lookup["Bodega"]),
                f"COPÉRNICO CSV fila {row_number}.Bodega",
                allow_none=True,
            )
            sku = to_id(
                row.get(field_lookup["EAN"]),
                f"COPÉRNICO CSV fila {row_number}.EAN",
                allow_none=True,
            )
            if warehouse is None or sku is None:
                continue
            source_rows += 1
            if copernico_is_usable(row.get(field_lookup["Ubicacion"])):
                continue
            unusable_rows += 1
            unusable_stock[(warehouse, sku)] += max(
                to_float(row.get(field_lookup["Saldo"]), 0.0),
                0.0,
            )

    warehouses = sorted({warehouse for warehouse, _ in unusable_stock})
    summary = {
        "total_rows": total_rows,
        "source_rows": source_rows,
        "unusable_rows": unusable_rows,
        "unusable_warehouse_skus": len(unusable_stock),
        "unusable_warehouses": warehouses,
        "unusable_units": sum(unusable_stock.values()),
    }
    return dict(unusable_stock), summary


@dataclass
class Catalogs:
    volume_m3: dict[int, float]
    blocked_products: set[int]
    route_cost_blocks: set[tuple[int, int]]
    store_priority: dict[int, int]
    # Clasificación referencial por (warehouse origen, SKU).
    high_value: dict[tuple[int, int], str]
    rackeados_444: set[int]
    store_capacity: dict[int, float]
    copernico_unusable_444: dict[int, float]
    unavailable_stock: dict[tuple[int, int], float]
    stock_base: dict[tuple[int, int], float]
    golden_infaltables: set[tuple[int, str]]
    stores: dict[int, dict[str, str]]
    storage: dict[int, str]
    warnings: list[str]
    storage_override_by_origin: dict[tuple[int, int], str] = field(
        default_factory=dict
    )
    origin_storage_override_enabled: bool = False
    copernico_unusable_by_warehouse: dict[tuple[int, int], float] = field(
        default_factory=dict
    )
    kvi_products: set[tuple[int, int]] = field(default_factory=set)
    excluded_products: set[int] = field(default_factory=set)
    infaltable_products: set[tuple[int, int]] = field(default_factory=set)
    golden_products: set[tuple[int, int]] = field(default_factory=set)
    anchor_products: set[tuple[int, int]] = field(default_factory=set)


def load_catalogs(
    path: Path,
    config: Config,
    *,
    copernico_csv_path: Path | None = None,
) -> Catalogs:
    warnings: list[str] = []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        volume_m3: dict[int, float] = {}
        for row in iter_sheet_records(
            workbook, "VOLUMETRIA", ["SKU", "PALLETS"]
        ):
            sku = to_id(row["SKU"], "VOLUMETRIA.SKU", allow_none=True)
            if sku is None:
                continue
            volume = to_float(row["PALLETS"], config.default_m3_per_unit)
            if volume <= 0:
                warnings.append(
                    f"VOLUMETRIA: SKU {sku} tiene volumen {volume}; se usó default "
                    f"{config.default_m3_per_unit}"
                )
                volume = config.default_m3_per_unit
            put_unique(volume_m3, sku, volume, warnings, "VOLUMETRIA", "first")

        blocked_products = {
            to_id(row["SKU"], "BLOQUEOS.SKU")
            for row in iter_sheet_records(workbook, "BLOQUEOS", ["SKU"])
        }

        route_cost_blocks: set[tuple[int, int]] = set()
        for row in iter_sheet_records(
            workbook, "RUTA_COSTOS", ["Destination", "Catalog ID"]
        ):
            destination = to_id(row["Destination"], "RUTA_COSTOS.Destination", True)
            sku = to_id(row["Catalog ID"], "RUTA_COSTOS.Catalog ID", True)
            if destination is not None and sku is not None:
                route_cost_blocks.add((destination, sku))

        store_priority: dict[int, int] = {}
        for row in iter_sheet_records(
            workbook, "PRIORIDAD", ["WAREHOUSE_ID", "PRIORIDAD"]
        ):
            warehouse = to_id(row["WAREHOUSE_ID"], "PRIORIDAD.WAREHOUSE_ID", True)
            if warehouse is None:
                continue
            priority = int(round(to_float(row["PRIORIDAD"], 100)))
            put_unique(store_priority, warehouse, priority, warnings, "PRIORIDAD", "min")

        high_value: dict[tuple[int, int], str] = {}
        for source, sheet_name in ((444, "444_HV"), (831, "831_HV")):
            for row in iter_sheet_records(
                workbook,
                sheet_name,
                ["EAN", "Category"],
            ):
                # En estas hojas EAN representa PRODUCT_ID.
                sku = to_id(row["EAN"], f"{sheet_name}.EAN", True)
                if sku is not None:
                    category = clean_text(row["Category"]) or "HV"
                    put_unique(
                        high_value,
                        (source, sku),
                        category,
                        warnings,
                        sheet_name,
                        "first",
                    )

        storage_override_by_origin: dict[tuple[int, int], str] = {}
        for row in iter_sheet_records(
            workbook,
            "OVER_ORIGEN_STORAGE",
            ["WAREHOUSE_ID", "PRODUCT_ID", "STORAGE_TYPE"],
        ):
            source = to_id(
                row["WAREHOUSE_ID"],
                "OVER_ORIGEN_STORAGE.WAREHOUSE_ID",
                True,
            )
            sku = to_id(
                row["PRODUCT_ID"],
                "OVER_ORIGEN_STORAGE.PRODUCT_ID",
                True,
            )
            storage_type = clean_text(row["STORAGE_TYPE"])
            if source is not None and sku is not None and storage_type:
                put_unique(
                    storage_override_by_origin,
                    (source, sku),
                    storage_type,
                    warnings,
                    "OVER_ORIGEN_STORAGE",
                    "first",
                )

        rackeados_444: set[int] = set()
        for row in iter_sheet_records(workbook, "RACKEADOS", ["WHS", "SYNC"]):
            warehouse = to_id(row["WHS"], "RACKEADOS.WHS", True)
            sku = to_id(row["SYNC"], "RACKEADOS.SYNC", True)
            if warehouse == 444 and sku is not None:
                rackeados_444.add(sku)

        store_capacity: dict[int, float] = {}
        for row in iter_sheet_records(workbook, "CAP_RECIBO", ["WH_ID", "CAP"]):
            warehouse = to_id(row["WH_ID"], "CAP_RECIBO.WH_ID", True)
            if warehouse is None:
                continue
            capacity = to_float(row["CAP"], config.default_store_capacity_m3)
            if capacity < 0:
                raise ValueError(f"CAP_RECIBO: capacidad negativa para warehouse {warehouse}")
            put_unique(
                store_capacity,
                warehouse,
                capacity,
                warnings,
                "CAP_RECIBO",
                "min",
            )

        if copernico_csv_path is not None:
            copernico_unusable_by_warehouse, copernico_summary = (
                load_copernico_unusable_csv(copernico_csv_path)
            )
            copernico_unusable_444 = {
                sku: quantity
                for (warehouse, sku), quantity in (
                    copernico_unusable_by_warehouse.items()
                )
                if warehouse == 444
            }
            warehouse_text = ", ".join(
                map(str, copernico_summary["unusable_warehouses"])
            ) or "sin bodegas con saldo no usable"
            warnings.append(
                "COPÉRNICO CSV: se procesaron "
                f"{copernico_summary['total_rows']:,} filas; "
                f"{copernico_summary['unusable_rows']:,} ubicaciones no usables de "
                f"{copernico_summary['unusable_warehouse_skus']:,} combinaciones "
                f"bodega-SKU descontaron "
                f"{copernico_summary['unusable_units']:,.0f} unidades. "
                f"Bodegas afectadas: {warehouse_text}."
            )
        else:
            copernico_unusable_444 = defaultdict(float)
            for row in iter_sheet_records(
                workbook,
                "COPERNICO",
                ["Bodega", "EAN", "Ubicacion", "Saldo"],
            ):
                warehouse = to_id(row["Bodega"], "COPERNICO.Bodega", True)
                sku = to_id(row["EAN"], "COPERNICO.EAN", True)
                if (
                    warehouse == 444
                    and sku is not None
                    and not copernico_is_usable(row["Ubicacion"])
                ):
                    copernico_unusable_444[sku] += max(
                        to_float(row["Saldo"], 0.0),
                        0.0,
                    )
            copernico_unusable_by_warehouse = {
                (444, sku): quantity
                for sku, quantity in copernico_unusable_444.items()
            }

        unavailable_stock: dict[tuple[int, int], float] = defaultdict(float)
        for row in iter_sheet_records(
            workbook,
            "NO_DISPONIBLE",
            ["WAREHOUSE_ID", "PRODUCT_ID", "STOCK"],
        ):
            warehouse = to_id(row["WAREHOUSE_ID"], "NO_DISPONIBLE.WAREHOUSE_ID", True)
            sku = to_id(row["PRODUCT_ID"], "NO_DISPONIBLE.PRODUCT_ID", True)
            if warehouse is not None and sku is not None:
                unavailable_stock[(warehouse, sku)] += max(to_float(row["STOCK"]), 0.0)

        stock_base: dict[tuple[int, int], float] = {}
        for row in iter_sheet_records(
            workbook,
            "STOCK",
            ["WAREHOUSE_ID", "PRODUCT_ID", "STOCK_DISPONIBLE_FINAL"],
        ):
            warehouse = to_id(row["WAREHOUSE_ID"], "STOCK.WAREHOUSE_ID", True)
            sku = to_id(row["PRODUCT_ID"], "STOCK.PRODUCT_ID", True)
            if warehouse is None or sku is None:
                continue
            stock = max(to_float(row["STOCK_DISPONIBLE_FINAL"]), 0.0)
            put_unique(stock_base, (warehouse, sku), stock, warnings, "STOCK", "min")

        # Campo legado conservado vacío para compatibilidad de objetos Catalogs.
        golden_infaltables: set[tuple[int, str]] = set()
        infaltable_products: set[tuple[int, int]] = set()
        golden_products: set[tuple[int, int]] = set()
        anchor_products: set[tuple[int, int]] = set()
        for row in iter_sheet_records(
            workbook,
            "GOLDEN_INFALTABLES_ANCHOR",
            [
                "WAREHOUSE_ID",
                "PRODUCT_ID_SYNC",
                "IS_INFALTABLE",
                "IS_GOLDEN",
                "IS_ANCHOR",
            ],
        ):
            destination = to_id(
                row["WAREHOUSE_ID"],
                "GOLDEN_INFALTABLES_ANCHOR.WAREHOUSE_ID",
                True,
            )
            sku = to_id(
                row["PRODUCT_ID_SYNC"],
                "GOLDEN_INFALTABLES_ANCHOR.PRODUCT_ID_SYNC",
                True,
            )
            if destination is None or sku is None:
                continue

            def active_flag(value: Any) -> bool:
                return clean_text(value).upper() in {
                    "1", "TRUE", "VERDADERO", "YES", "SI", "SÍ", "Y",
                }

            key = (destination, sku)
            if active_flag(row["IS_INFALTABLE"]):
                infaltable_products.add(key)
            if active_flag(row["IS_GOLDEN"]):
                golden_products.add(key)
            if active_flag(row["IS_ANCHOR"]):
                anchor_products.add(key)

        kvi_products: set[tuple[int, int]] = set()
        for row in iter_sheet_records(
            workbook,
            "KVI",
            ["WAREHOUSE_ID", "PRODUCT_ID", "KVI"],
        ):
            warehouse = to_id(row["WAREHOUSE_ID"], "KVI.WAREHOUSE_ID", True)
            sku = to_id(row["PRODUCT_ID"], "KVI.PRODUCT_ID", True)
            kvi_flag = clean_text(row["KVI"]).upper()
            is_kvi = kvi_flag not in {"0", "FALSE", "FALSO", "NO", "N"}
            if warehouse is not None and sku is not None and is_kvi:
                kvi_products.add((warehouse, sku))

        stores: dict[int, dict[str, str]] = {}
        for row in iter_sheet_records(
            workbook,
            "TIENDA",
            ["CITY", "WAREHOUSE_ID", "WAREHOUSE_NAME"],
        ):
            warehouse = to_id(row["WAREHOUSE_ID"], "TIENDA.WAREHOUSE_ID", True)
            if warehouse is None:
                continue
            value = {
                "city": clean_text(row["CITY"]),
                "city_norm": normalize_city(row["CITY"]),
                "warehouse_name": clean_text(row["WAREHOUSE_NAME"]),
            }
            put_unique(stores, warehouse, value, warnings, "TIENDA", "error")

        storage: dict[int, str] = {}
        for row in iter_sheet_records(
            workbook,
            "STORAGE",
            ["PRODUCT_ID", "STORAGE_NAME"],
        ):
            sku = to_id(row["PRODUCT_ID"], "STORAGE.PRODUCT_ID", True)
            if sku is not None:
                value = clean_text(row["STORAGE_NAME"]) or "UNKNOWN"
                put_unique(storage, sku, value, warnings, "STORAGE", "first")
    finally:
        workbook.close()

    for origin in config.origin_warehouses:
        if origin not in stores:
            raise ValueError(
                f"El origen {origin} no existe en TIENDA; no puedo evaluar bloqueos regionales"
            )

    return Catalogs(
        volume_m3=volume_m3,
        blocked_products=blocked_products,
        route_cost_blocks=route_cost_blocks,
        store_priority=store_priority,
        high_value=high_value,
        rackeados_444=rackeados_444,
        store_capacity=store_capacity,
        copernico_unusable_444=dict(copernico_unusable_444),
        unavailable_stock=dict(unavailable_stock),
        stock_base=stock_base,
        golden_infaltables=golden_infaltables,
        stores=stores,
        storage=storage,
        warnings=warnings,
        copernico_unusable_by_warehouse=dict(
            copernico_unusable_by_warehouse
        ),
        kvi_products=kvi_products,
        storage_override_by_origin=storage_override_by_origin,
        infaltable_products=infaltable_products,
        golden_products=golden_products,
        anchor_products=anchor_products,
    )


def source_value_category(catalogs: Catalogs, source: int, sku: int) -> str:
    """Devuelve la categoría de valor aplicable al origen real de la línea."""
    return catalogs.high_value.get((source, sku), "REGULAR")


def source_storage_type(catalogs: Catalogs, source: int, sku: int) -> str:
    """Resuelve STORAGE por origen cuando el override está habilitado."""
    if catalogs.origin_storage_override_enabled:
        override = clean_text(
            catalogs.storage_override_by_origin.get((source, sku))
        )
        if override:
            return override
    storage_type = clean_text(catalogs.storage.get(sku))
    if not storage_type or storage_type.upper() in {"UNKNOWN", "N/A", "NULL"}:
        return "Room Temperature"
    return storage_type


def product_priority_profile(
    catalogs: Catalogs,
    destination: int,
    sku: int,
) -> dict[str, Any]:
    """Clasifica tienda–SKU con jerarquía INFALTABLE > GOLDEN > ANCHOR."""
    key = (destination, sku)
    is_infaltable = key in catalogs.infaltable_products
    is_golden = key in catalogs.golden_products
    is_anchor = key in catalogs.anchor_products
    if is_infaltable:
        product_type, rank = "INFALTABLE", 0
    elif is_golden:
        product_type, rank = "GOLDEN", 1
    elif is_anchor:
        product_type, rank = "ANCHOR", 2
    elif key in catalogs.kvi_products:
        product_type, rank = "KVI", 3
    else:
        product_type, rank = "REGULAR", 4
    return {
        "is_infaltable": is_infaltable,
        "is_golden": is_golden,
        "is_anchor": is_anchor,
        "is_kvi": key in catalogs.kvi_products,
        "type": product_type,
        "rank": rank,
    }


def allocation_value_summary(
    catalogs: Catalogs,
    allocations: Iterable[tuple[int, int]],
    sku: int,
) -> str:
    """Resume VALUE en una fila de reporte que puede utilizar varios orígenes."""
    used_sources = [source for source, quantity in allocations if quantity > 0]
    if not used_sources:
        return "REGULAR"
    values = [source_value_category(catalogs, source, sku) for source in used_sources]
    if len(set(values)) == 1:
        return values[0]
    return " | ".join(
        f"{source}:{source_value_category(catalogs, source, sku)}"
        for source in used_sources
    )


def allocation_storage_summary(
    catalogs: Catalogs,
    allocations: Iterable[tuple[int, int]],
    sku: int,
) -> str:
    """Resume STORAGE para un caso atendido desde uno o varios orígenes."""
    used_sources = [source for source, quantity in allocations if quantity > 0]
    if not used_sources:
        return source_storage_type(catalogs, 0, sku)
    values = [source_storage_type(catalogs, source, sku) for source in used_sources]
    if len(set(values)) == 1:
        return values[0]
    return " | ".join(
        f"{source}:{source_storage_type(catalogs, source, sku)}"
        for source in used_sources
    )


def detect_fountain9_store_outliers(
    consolidated: dict[tuple[int, int], dict[str, Any]],
    catalogs: Catalogs,
) -> dict[str, Any]:
    """Detecta tiendas con una cobertura de líneas F9 anormalmente baja."""
    line_counts = Counter(destination for destination, _ in consolidated)
    median_lines = float(statistics.median(line_counts.values())) if line_counts else 0.0
    threshold = int(math.floor(median_lines * 0.50))
    detection_enabled = len(line_counts) >= 5 and median_lines >= 20
    outlier_store_ids = {
        destination
        for destination, count in line_counts.items()
        if detection_enabled and count <= threshold
    }
    stores: list[dict[str, Any]] = []
    for destination in sorted(outlier_store_ids, key=lambda value: (line_counts[value], value)):
        store = catalogs.stores.get(destination, {})
        count = int(line_counts[destination])
        stores.append(
            {
                "WAREHOUSE_DESTINATION": destination,
                "WAREHOUSE_NAME": store.get("warehouse_name", ""),
                "CITY": store.get("city", ""),
                "LINEAS_F9_UNICAS": count,
                "MEDIANA_LINEAS_TIENDAS": round(median_lines, 1),
                "UMBRAL_OUTLIER_50_PCT": threshold,
                "PORCENTAJE_DE_LA_MEDIANA": round(
                    (count / median_lines) * 100,
                    1,
                ) if median_lines else 0.0,
                "ACCION": "EXCLUIDA DE LA PLANEACION",
            }
        )
    details: list[dict[str, Any]] = []
    for (destination, sku), record in sorted(consolidated.items()):
        if destination not in outlier_store_ids:
            continue
        store = catalogs.stores.get(destination, {})
        details.append(
            {
                "WAREHOUSE_DESTINATION": destination,
                "WAREHOUSE_NAME": store.get("warehouse_name", ""),
                "CITY": store.get("city", ""),
                "RETAIL_ID": sku,
                "PREDICTED_DEMAND": record["PREDICTED_DEMAND"],
                "PREDICTED_OPENING_INVENTORY": record[
                    "PREDICTED_OPENING_INVENTORY"
                ],
                "ROQ_INPUT": record["ROQ_INPUT"],
                "NET_INTER_STORE_TRANSFERS": record[
                    "NET_INTER_STORE_TRANSFERS"
                ],
                "ARCHIVOS_INPUT": " | ".join(sorted(record["SOURCE_FILES"])),
                "FILAS_INPUT_SUMADAS": record["SOURCE_ROWS"],
                "MOTIVO_EXCLUSION": (
                    f"La tienda solo contiene {line_counts[destination]} líneas "
                    f"únicas frente a una mediana de {median_lines:g}; umbral "
                    f"automático: {threshold}."
                ),
            }
        )
    return {
        "enabled": detection_enabled,
        "stores_evaluated": len(line_counts),
        "median_lines": median_lines,
        "threshold": threshold,
        "store_ids": sorted(outlier_store_ids),
        "stores": stores,
        "details": details,
        "stores_excluded": len(outlier_store_ids),
        "requirements_excluded": len(details),
    }


# %% Lectura y preparación del requerimiento diario

PLAN_REQUIRED_COLUMNS = [
    "Warehouseid",
    "SKU ID",
    "Current Inventory",
    "Predicted Demand for selected duration",
    "Predicted Opening Inventory",
    "Replenishment Quantity for Plan Duration (MOV)",
]


@dataclass
class PlanReadResult:
    rows: list[dict[str, Any]]
    warnings: list[str]
    input_row_count: int
    duplicate_row_count: int
    duplicate_key_count: int
    conflicting_duplicate_key_count: int


def duplicate_selection_score(row: dict[str, Any], config: Config) -> tuple[Any, ...]:
    """Escoge la versión más exigente sin sumar dos veces el mismo requerimiento."""
    target, _ = calculate_target_quantity(row, config)
    shortfall = row["PREDICTED_DEMAND"] - row["PREDICTED_OPENING_INVENTORY"]
    return (
        target,
        row["MOV_ORIGINAL"],
        shortfall,
        row["PREDICTED_DEMAND"],
        -row["PREDICTED_OPENING_INVENTORY"],
        -row["INPUT_ROW"],
    )


def read_plan_csv(path: Path, config: Config) -> PlanReadResult:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        if reader.fieldnames is None:
            raise ValueError("El CSV diario no contiene encabezados")
        field_lookup = {normalize_header(name): name for name in reader.fieldnames}
        missing = [
            name for name in PLAN_REQUIRED_COLUMNS if normalize_header(name) not in field_lookup
        ]
        if missing:
            raise ValueError(f"El CSV diario no contiene las columnas obligatorias: {missing}")

        consolidated: dict[tuple[int, int], dict[str, Any]] = {}
        input_row_count = 0
        for input_row, raw in enumerate(reader, start=2):
            input_row_count += 1
            destination = to_id(
                raw[field_lookup[normalize_header("Warehouseid")]],
                f"CSV fila {input_row}.Warehouseid",
            )
            sku = to_id(
                raw[field_lookup[normalize_header("SKU ID")]],
                f"CSV fila {input_row}.SKU ID",
            )
            key = (destination, sku)

            def get(column: str) -> Any:
                return raw.get(field_lookup.get(normalize_header(column), ""), "")

            candidate = {
                "INPUT_ROW": input_row,
                "WAREHOUSE_DESTINATION": destination,
                "RETAIL_ID": sku,
                "SKU_NAME": clean_text(get("SKU Name")),
                "CSV_WAREHOUSE_NAME": clean_text(get("Warehouse Name")),
                "CURRENT_INVENTORY": to_float(get("Current Inventory")),
                "PREDICTED_DEMAND": to_float(
                    get("Predicted Demand for selected duration")
                ),
                "PREDICTED_OPENING_INVENTORY": to_float(
                    get("Predicted Opening Inventory")
                ),
                "MOV_ORIGINAL": to_float(
                    get("Replenishment Quantity for Plan Duration (MOV)")
                ),
            }
            signature = (
                candidate["CURRENT_INVENTORY"],
                candidate["PREDICTED_DEMAND"],
                candidate["PREDICTED_OPENING_INVENTORY"],
                candidate["MOV_ORIGINAL"],
            )
            if key not in consolidated:
                consolidated[key] = {
                    "chosen": candidate,
                    "count": 1,
                    "row_numbers": [input_row],
                    "signatures": {signature},
                }
                continue

            entry = consolidated[key]
            entry["count"] += 1
            if len(entry["row_numbers"]) < 20:
                entry["row_numbers"].append(input_row)
            entry["signatures"].add(signature)
            if duplicate_selection_score(candidate, config) > duplicate_selection_score(
                entry["chosen"], config
            ):
                entry["chosen"] = candidate

    rows: list[dict[str, Any]] = []
    duplicate_key_count = 0
    conflicting_key_count = 0
    conflicting_examples: list[str] = []
    for key, entry in consolidated.items():
        record = dict(entry["chosen"])
        count = int(entry["count"])
        conflict = len(entry["signatures"]) > 1
        if count > 1:
            duplicate_key_count += 1
        if conflict:
            conflicting_key_count += 1
            if len(conflicting_examples) < 10:
                conflicting_examples.append(str(key))
        displayed_rows = ",".join(str(value) for value in entry["row_numbers"])
        if count > len(entry["row_numbers"]):
            displayed_rows += ",..."
        record.update(
            {
                "INPUT_DUPLICATE_COUNT": count,
                "INPUT_ROWS_MERGED": displayed_rows,
                "INPUT_DUPLICATE_CONFLICT": conflict,
            }
        )
        rows.append(record)

    duplicate_row_count = input_row_count - len(rows)
    warnings: list[str] = []
    if duplicate_row_count:
        warning = (
            f"CSV: se consolidaron {duplicate_row_count:,} filas repetidas. "
            f"Llaves destino-SKU afectadas: {duplicate_key_count:,}. No se sumaron: se eligió "
            "la fila que genera la mayor cantidad objetivo."
        )
        if conflicting_key_count:
            warning += (
                f" Llaves con métricas distintas: {conflicting_key_count:,}; "
                f"ejemplos: {', '.join(conflicting_examples)}."
            )
        warnings.append(warning)

    return PlanReadResult(
        rows=rows,
        warnings=warnings,
        input_row_count=input_row_count,
        duplicate_row_count=duplicate_row_count,
        duplicate_key_count=duplicate_key_count,
        conflicting_duplicate_key_count=conflicting_key_count,
    )


def calculate_target_quantity(row: dict[str, Any], config: Config) -> tuple[int, str]:
    mov = row["MOV_ORIGINAL"]
    demand = row["PREDICTED_DEMAND"]
    opening = row["PREDICTED_OPENING_INVENTORY"]
    if mov > 0:
        return max(int(math.ceil(mov)), config.minimum_positive_quantity), "MOV_MINIMO_3"
    if math.isclose(demand, 0.0, abs_tol=1e-9) and math.isclose(
        opening, 0.0, abs_tol=1e-9
    ):
        return 4, "HARDCODE_4_CERO_TOTAL"
    if opening < demand:
        return 3, "HARDCODE_3_INVENTARIO_MENOR_DEMANDA"
    return 0, "SIN_DEMANDA"


# %% Motor de asignación

def source_stock_components(catalogs: Catalogs, source: int, sku: int) -> dict[str, Any]:
    base = max(catalogs.stock_base.get((source, sku), 0.0), 0.0)
    unavailable = max(catalogs.unavailable_stock.get((source, sku), 0.0), 0.0)
    copernico_unusable = max(
        catalogs.copernico_unusable_by_warehouse.get(
            (source, sku),
            catalogs.copernico_unusable_444.get(sku, 0.0)
            if source == 444
            else 0.0,
        ),
        0.0,
    )
    rackeado = source == 444 and sku in catalogs.rackeados_444
    adjusted = max(base - unavailable - copernico_unusable, 0.0)
    if rackeado or sku in catalogs.excluded_products:
        adjusted = 0.0
    return {
        "base": base,
        "unavailable": unavailable,
        "copernico_unusable": copernico_unusable,
        "rackeado": rackeado,
        "adjusted": int(math.floor(adjusted + 1e-9)),
    }


def is_regional_block(
    catalogs: Catalogs,
    source: int,
    destination: int,
    sku: int,
    destination_city_norm: str,
    is_golden_infaltable: bool,
) -> bool:
    """Aplica únicamente los bloqueos explícitos de la hoja BLOQUEOS.

    ``is_golden_infaltable`` se conserva en la firma por compatibilidad con los
    distintos engines, pero Golden Infaltables ya no genera una restricción de
    origen o ciudad. Su clasificación sigue disponible para prioridad y
    reporting.
    """
    if source == destination:
        return True
    source_city_norm = catalogs.stores[source]["city_norm"]
    restricted_product = sku in catalogs.blocked_products
    return (
        restricted_product
        and source_city_norm == CDMX
        and destination_city_norm in FOREIGN_DESTINATION_CITIES
    )


@dataclass
class PlanningResult:
    base_rows: list[dict[str, Any]]
    allocation_rows: list[dict[str, Any]]
    capacity_rows: list[dict[str, Any]]
    warnings: list[str]
    tasks_used: int
    max_tasks: int


def plan_transfers(
    plan_rows: list[dict[str, Any]],
    catalogs: Catalogs,
    config: Config,
) -> PlanningResult:
    prepared: list[dict[str, Any]] = []
    for row in plan_rows:
        destination = row["WAREHOUSE_DESTINATION"]
        sku = row["RETAIL_ID"]
        if sku in catalogs.excluded_products:
            continue
        store = catalogs.stores.get(destination)
        city = store["city"] if store else ""
        city_norm = store["city_norm"] if store else ""
        warehouse_name = (
            store["warehouse_name"]
            if store
            else row.get("CSV_WAREHOUSE_NAME", "")
        )
        target, demand_rule = calculate_target_quantity(row, config)
        priority_profile = product_priority_profile(catalogs, destination, sku)
        is_infaltable = priority_profile["is_infaltable"]
        is_golden = priority_profile["is_golden"]
        is_anchor = priority_profile["is_anchor"]
        is_kvi = priority_profile["is_kvi"]
        enriched = dict(row)
        enriched.update(
            {
                "CITY": city,
                "CITY_NORMALIZED": city_norm,
                "WAREHOUSE_NAME": warehouse_name,
                "PRIORIDAD_TIENDA": catalogs.store_priority.get(destination, 100),
                "ES_MANUAL_FORECAST_ZERO": bool(
                    row.get("ES_MANUAL_FORECAST_ZERO", False)
                ),
                "ES_INFALTABLE": is_infaltable,
                "ES_GOLDEN": is_golden,
                "ES_ANCHOR": is_anchor,
                "TIPO_PRIORIDAD_PRODUCTO": priority_profile["type"],
                "RANGO_PRIORIDAD_PRODUCTO": priority_profile["rank"],
                "ES_GOLDEN_INFALTABLE": is_infaltable or is_golden,
                "ES_KVI": is_kvi,
                "ES_STOCKOUT": row["CURRENT_INVENTORY"] <= 0,
                "M3_POR_UNIDAD": catalogs.volume_m3.get(
                    sku, config.default_m3_per_unit
                ),
                "STORAGE": catalogs.storage.get(sku, "UNKNOWN"),
                "VALUE": "POR ORIGEN",
                "CANTIDAD_OBJETIVO": target,
                "REGLA_DEMANDA": demand_rule,
                "SIN_RUTA_COSTOS": (destination, sku)
                in catalogs.route_cost_blocks,
            }
        )
        prepared.append(enriched)

    prepared.sort(
        key=lambda row: (
            row["RANGO_PRIORIDAD_PRODUCTO"],
            0 if not row["ES_MANUAL_FORECAST_ZERO"] else 1,
            row["PRIORIDAD_TIENDA"],
            0 if row["ES_STOCKOUT"] else 1,
            row["WAREHOUSE_DESTINATION"],
            row["RETAIL_ID"],
            row["INPUT_ROW"],
        )
    )

    stock_info_cache: dict[tuple[int, int], dict[str, Any]] = {}
    stock_remaining: dict[tuple[int, int], int] = {}

    def get_stock_info(source: int, sku: int) -> dict[str, Any]:
        key = (source, sku)
        if key not in stock_info_cache:
            info = source_stock_components(catalogs, source, sku)
            stock_info_cache[key] = info
            stock_remaining[key] = info["adjusted"]
        return stock_info_cache[key]

    cap_used_normal: dict[int, float] = defaultdict(float)
    actual_m3_by_store: dict[int, float] = defaultdict(float)
    cap_closed: dict[int, bool] = defaultdict(bool)
    cap_crossed_by_line: dict[int, bool] = defaultdict(bool)
    tasks_used = 0
    base_rows: list[dict[str, Any]] = []
    allocation_rows: list[dict[str, Any]] = []
    deferred_stock_rows: list[tuple[dict[str, Any], dict[str, Any]]] = []

    for planning_order, row in enumerate(prepared, start=1):
        destination = row["WAREHOUSE_DESTINATION"]
        sku = row["RETAIL_ID"]
        target = row["CANTIDAD_OBJETIVO"]
        city = row["CITY"]
        city_norm = row["CITY_NORMALIZED"]
        is_golden = row["ES_GOLDEN"]
        m3_per_unit = row["M3_POR_UNIDAD"]
        capacity = catalogs.store_capacity.get(
            destination, config.default_store_capacity_m3
        )
        cap_before = cap_used_normal[destination]
        task_before = tasks_used

        origin_info: dict[int, dict[str, Any]] = {}
        origin_before: dict[int, int] = {}
        regional_blocks: dict[int, bool] = {}
        for source in config.origin_warehouses:
            info = get_stock_info(source, sku)
            origin_info[source] = info
            origin_before[source] = stock_remaining[(source, sku)]
            regional_blocks[source] = is_regional_block(
                catalogs,
                source,
                destination,
                sku,
                city_norm,
                is_golden,
            )

        allocations: list[tuple[int, int]] = []
        candidate_allocations: list[tuple[int, int]] = []
        tipo_corte = ""
        detalle_motivo = ""
        passes_capacity = True
        passes_tasks = True
        stock_insufficient = False
        capacity_target = 0

        if target <= 0:
            tipo_corte = "SIN DEMANDA"
            detalle_motivo = "MOV cero y no activa reglas de hardcode"
        elif not city_norm:
            tipo_corte = "ERROR DE DATOS"
            detalle_motivo = "El warehouse destino no existe en TIENDA"
            passes_capacity = False
            passes_tasks = False
        elif row["SIN_RUTA_COSTOS"]:
            tipo_corte = "SIN RUTA DE COSTOS"
            detalle_motivo = "Destino-SKU listado en RUTA_COSTOS"
            passes_capacity = False
            passes_tasks = False
        elif cap_closed[destination]:
            tipo_corte = "CORTE POR CAPACIDAD DE TIENDA"
            detalle_motivo = (
                f"La tienda alcanzó {capacity:.6f} m3 en una línea anterior"
            )
            passes_capacity = False
        else:
            capacity_remaining_m3 = max(capacity - cap_before, 0.0)
            if m3_per_unit > 0:
                capacity_units = int(
                    math.floor((capacity_remaining_m3 / m3_per_unit) + 1e-9)
                )
            else:
                capacity_units = target
            capacity_target = min(target, max(capacity_units, 0))
            capacity_limited = capacity_target < target
            remaining_candidate = capacity_target
            for source in config.origin_warehouses:
                if regional_blocks[source]:
                    continue
                available = stock_remaining[(source, sku)]
                quantity = min(remaining_candidate, available)
                if quantity > 0:
                    candidate_allocations.append((source, quantity))
                    remaining_candidate -= quantity
                if remaining_candidate <= 0:
                    break

            # Una necesidad que no puede cubrirse completa con el stock
            # elegible remanente no debe consumir inventario. Se omite esta
            # tienda y se conserva el saldo para requerimientos posteriores
            # que sí puedan atenderse al 100%.
            stock_insufficient = remaining_candidate > 0
            if stock_insufficient:
                candidate_allocations = []

            available_task_slots = max(config.max_tasks - tasks_used, 0)
            allocations = candidate_allocations[:available_task_slots]
            passes_tasks = len(allocations) == len(candidate_allocations)

            for source, quantity in allocations:
                stock_remaining[(source, sku)] -= quantity
                tasks_used += 1
                allocation_rows.append(
                    {
                        "WAREHOUSE_DESTINATION": destination,
                        "WAREHOUSE_SOURCE": source,
                        "RETAIL_ID": sku,
                        "QUANTITY": int(quantity),
                        "PLANNED_DATE": "",
                        "ROUTE": 1,
                        "DELIVERY_PRIORITY": 1,
                        "CITY": city,
                        "STORAGE": source_storage_type(catalogs, source, sku),
                        "VALUE": source_value_category(catalogs, source, sku),
                    }
                )

            assigned = sum(quantity for _, quantity in allocations)
            candidate_total = sum(quantity for _, quantity in candidate_allocations)
            missing = target - assigned
            blocked_stock = sum(
                origin_before[source]
                for source in config.origin_warehouses
                if regional_blocks[source]
            )

            if assigned > 0:
                actual_m3 = assigned * m3_per_unit
                actual_m3_by_store[destination] += actual_m3
                cap_used_normal[destination] += actual_m3
                if cap_used_normal[destination] >= capacity - 1e-9:
                    cap_closed[destination] = True

            if assigned >= target:
                tipo_corte = "OK"
                detalle_motivo = "Demanda cubierta completamente"
            elif assigned > 0 and not passes_tasks:
                tipo_corte = "OK PARCIAL - CORTE POR CAPACIDAD DE TAREAS"
                detalle_motivo = f"Asignadas {assigned} de {target}; límite global de tareas"
            elif (
                assigned > 0
                and capacity_limited
                and assigned >= capacity_target
            ):
                tipo_corte = "OK PARCIAL - CORTE POR CAPACIDAD DE TIENDA"
                detalle_motivo = (
                    f"Asignadas {assigned} de {target}; se agotó la capacidad "
                    f"restante de {capacity_remaining_m3:.6f} m3"
                )
                passes_capacity = False
            elif assigned > 0 and blocked_stock > 0:
                tipo_corte = "OK PARCIAL - BLOQUEO REGIONAL"
                detalle_motivo = f"Asignadas {assigned} de {target}; stock bloqueado CDMX→GDL/MTY"
            elif assigned > 0:
                tipo_corte = "OK PARCIAL - CORTE POR STOCK"
                detalle_motivo = f"Asignadas {assigned} de {target}; stock permitido insuficiente"
            elif candidate_total > 0 and not passes_tasks:
                tipo_corte = "CORTE POR CAPACIDAD DE TAREAS"
                detalle_motivo = "No quedaron tareas disponibles"
            elif capacity_limited and capacity_target <= 0:
                tipo_corte = "CORTE POR CAPACIDAD DE TIENDA"
                detalle_motivo = "No queda capacidad de recibo para una unidad adicional"
                passes_capacity = False
            elif blocked_stock > 0:
                tipo_corte = "BLOQUEO REGIONAL"
                detalle_motivo = "El stock disponible está en un origen CDMX bloqueado para GDL/MTY"
            else:
                tipo_corte = "CORTE POR STOCK"
                diagnostics: list[str] = []
                if origin_info.get(444, {}).get("rackeado"):
                    diagnostics.append("RACKEADO_444")
                if any(
                    info.get("copernico_unusable", 0) > 0
                    for info in origin_info.values()
                ):
                    diagnostics.append("COPERNICO_NO_USABLE")
                if any(info["unavailable"] > 0 for info in origin_info.values()):
                    diagnostics.append("NO_DISPONIBLE")
                suffix = f" ({', '.join(diagnostics)})" if diagnostics else ""
                eligible_remaining = sum(
                    origin_before[source]
                    for source in config.origin_warehouses
                    if not regional_blocks[source]
                )
                if stock_insufficient and eligible_remaining > 0:
                    detalle_motivo = (
                        f"Stock elegible remanente {eligible_remaining} menor que "
                        f"la cantidad requerida {capacity_target}; no se consumió "
                        "para permitir atender prioridades posteriores"
                        + suffix
                    )
                else:
                    detalle_motivo = (
                        "Sin stock permitido en los orígenes configurados" + suffix
                    )

        assigned_by_origin = {source: 0 for source in config.origin_warehouses}
        for source, quantity in allocations:
            assigned_by_origin[source] += quantity
        assigned_total = sum(assigned_by_origin.values())
        m3_assigned = assigned_total * m3_per_unit
        cap_after = cap_used_normal[destination]

        if "CAPACIDAD DE TIENDA" in tipo_corte:
            passes_capacity = False

        report_row: dict[str, Any] = {
            "ORDEN_PLANIFICACION": planning_order,
            "FILA_INPUT": row["INPUT_ROW"],
            "FILAS_INPUT_CONSOLIDADAS": row.get("INPUT_ROWS_MERGED", str(row["INPUT_ROW"])),
            "CANTIDAD_FILAS_INPUT": row.get("INPUT_DUPLICATE_COUNT", 1),
            "DUPLICADO_CONFLICTIVO": row.get("INPUT_DUPLICATE_CONFLICT", False),
            "WAREHOUSE_DESTINATION": destination,
            "WAREHOUSE_NAME": row["WAREHOUSE_NAME"],
            "CITY": city,
            "RETAIL_ID": sku,
            "SKU_NAME": row["SKU_NAME"],
            "PREDICTED_OPENING_INVENTORY": row["PREDICTED_OPENING_INVENTORY"],
            "PREDICTED_DEMAND": row["PREDICTED_DEMAND"],
            "CURRENT_INVENTORY": row["CURRENT_INVENTORY"],
            "MOV_ORIGINAL": row["MOV_ORIGINAL"],
            "REGLA_DEMANDA": row["REGLA_DEMANDA"],
            "CANTIDAD_OBJETIVO": target,
            "CANTIDAD_ASIGNADA": assigned_total,
            "CANTIDAD_FALTANTE": max(target - assigned_total, 0),
            "ES_INFALTABLE": row["ES_INFALTABLE"],
            "ES_GOLDEN": row["ES_GOLDEN"],
            "ES_ANCHOR": row["ES_ANCHOR"],
            "TIPO_PRIORIDAD_PRODUCTO": row["TIPO_PRIORIDAD_PRODUCTO"],
            "RANGO_PRIORIDAD_PRODUCTO": row["RANGO_PRIORIDAD_PRODUCTO"],
            "ES_GOLDEN_INFALTABLE": row["ES_GOLDEN_INFALTABLE"],
            "ES_KVI": row["ES_KVI"],
            "PRIORIDAD_TIENDA": row["PRIORIDAD_TIENDA"],
            "ES_STOCKOUT": row["ES_STOCKOUT"],
            "SIN_RUTA_COSTOS": row["SIN_RUTA_COSTOS"],
            "M3_POR_UNIDAD": m3_per_unit,
            "M3_OBJETIVO": target * m3_per_unit,
            "M3_ASIGNADO": m3_assigned,
            "CAPACIDAD_TIENDA_M3": capacity,
            "M3_CAPACIDAD_ANTES": cap_before,
            "M3_CAPACIDAD_DESPUES": cap_after,
            "EXCEDE_CAPACIDAD_EN_ESTA_LINEA": (
                assigned_total > 0 and cap_after > capacity + 1e-9
            ),
            "PASA_CAPACIDAD": passes_capacity,
            "TAREAS_ANTES": task_before,
            "TAREAS_GENERADAS": len(allocations),
            "TAREAS_ACUMULADAS": tasks_used,
            "PASA_TAREAS": passes_tasks,
            "ORIGENES_USADOS": " | ".join(
                f"{source}:{quantity}" for source, quantity in allocations
            ),
            "STORAGE": allocation_storage_summary(catalogs, allocations, sku),
            "VALUE": allocation_value_summary(catalogs, allocations, sku),
            "TIPO_DE_CORTE": tipo_corte,
            "DETALLE_MOTIVO": detalle_motivo,
        }

        for source in config.origin_warehouses:
            info = origin_info[source]
            report_row.update(
                {
                    f"STOCK_BASE_{source}": info["base"],
                    f"NO_DISPONIBLE_{source}": info["unavailable"],
                    f"COPERNICO_NO_USABLE_{source}": info["copernico_unusable"],
                    f"RACKEADO_{source}": info["rackeado"],
                    f"STOCK_INICIAL_AJUSTADO_{source}": info["adjusted"],
                    f"STOCK_ANTES_{source}": origin_before[source],
                    f"BLOQUEO_REGIONAL_{source}": regional_blocks[source],
                    f"VALUE_{source}": source_value_category(catalogs, source, sku),
                    f"STORAGE_{source}": source_storage_type(catalogs, source, sku),
                    f"ASIGNADO_{source}": assigned_by_origin[source],
                    f"STOCK_REMANENTE_{source}": stock_remaining[(source, sku)],
                }
            )
        base_rows.append(report_row)

        if stock_insufficient and assigned_total == 0 and capacity_target > 0:
            deferred_stock_rows.append((row, report_row))

    # Segunda pasada: una vez atendidas todas las necesidades que cabían
    # completas, el stock todavía remanente vuelve a las tiendas diferidas en
    # su orden original de prioridad. En esta etapa sí se permite una atención
    # parcial para aprovechar el saldo sin bloquear tiendas posteriores.
    for row, report_row in deferred_stock_rows:
        destination = row["WAREHOUSE_DESTINATION"]
        sku = row["RETAIL_ID"]
        target = row["CANTIDAD_OBJETIVO"]
        m3_per_unit = row["M3_POR_UNIDAD"]
        capacity = catalogs.store_capacity.get(
            destination, config.default_store_capacity_m3
        )
        capacity_before = cap_used_normal[destination]
        capacity_remaining_m3 = max(capacity - capacity_before, 0.0)
        if m3_per_unit > 0:
            capacity_units = int(
                math.floor((capacity_remaining_m3 / m3_per_unit) + 1e-9)
            )
        else:
            capacity_units = target
        second_pass_target = min(target, max(capacity_units, 0))
        if second_pass_target <= 0:
            continue

        second_pass_before: dict[int, int] = {}
        second_pass_blocks: dict[int, bool] = {}
        candidates: list[tuple[int, int]] = []
        remaining = second_pass_target
        for source in config.origin_warehouses:
            second_pass_before[source] = stock_remaining[(source, sku)]
            blocked = is_regional_block(
                catalogs,
                source,
                destination,
                sku,
                row["CITY_NORMALIZED"],
                row["ES_GOLDEN"],
            )
            second_pass_blocks[source] = blocked
            if blocked:
                continue
            quantity = min(remaining, stock_remaining[(source, sku)])
            if quantity > 0:
                candidates.append((source, quantity))
                remaining -= quantity

        available_task_slots = max(config.max_tasks - tasks_used, 0)
        second_allocations = candidates[:available_task_slots]
        if not second_allocations:
            continue

        task_before_second_pass = tasks_used
        assigned_by_origin = {
            source: 0 for source in config.origin_warehouses
        }
        for source, quantity in second_allocations:
            stock_remaining[(source, sku)] -= quantity
            assigned_by_origin[source] += quantity
            tasks_used += 1
            allocation_rows.append(
                {
                    "WAREHOUSE_DESTINATION": destination,
                    "WAREHOUSE_SOURCE": source,
                    "RETAIL_ID": sku,
                    "QUANTITY": int(quantity),
                    "PLANNED_DATE": "",
                    "ROUTE": 1,
                    "DELIVERY_PRIORITY": 1,
                    "CITY": row["CITY"],
                    "STORAGE": source_storage_type(catalogs, source, sku),
                    "VALUE": source_value_category(catalogs, source, sku),
                }
            )

        assigned = sum(quantity for _, quantity in second_allocations)
        assigned_m3 = assigned * m3_per_unit
        cap_used_normal[destination] += assigned_m3
        actual_m3_by_store[destination] += assigned_m3
        capacity_after = cap_used_normal[destination]
        if capacity_after >= capacity - 1e-9:
            cap_closed[destination] = True

        task_limited = len(second_allocations) < len(candidates)
        capacity_limited = second_pass_target < target
        blocked_stock = sum(
            second_pass_before[source]
            for source in config.origin_warehouses
            if second_pass_blocks[source]
        )
        if task_limited:
            tipo_corte = "OK PARCIAL - CORTE POR CAPACIDAD DE TAREAS"
            detalle_motivo = (
                f"Segunda pasada: asignadas {assigned} de {target}; "
                "se alcanzó el límite global de tareas"
            )
        elif capacity_limited and assigned >= second_pass_target:
            tipo_corte = "OK PARCIAL - CORTE POR CAPACIDAD DE TIENDA"
            detalle_motivo = (
                f"Segunda pasada: asignadas {assigned} de {target}; "
                "se agotó la capacidad restante de la tienda"
            )
        elif blocked_stock > 0:
            tipo_corte = "OK PARCIAL - BLOQUEO REGIONAL"
            detalle_motivo = (
                f"Segunda pasada: asignadas {assigned} de {target}; parte del "
                "stock permanece bloqueada por la regla regional explícita"
            )
        else:
            tipo_corte = "OK PARCIAL - CORTE POR STOCK"
            detalle_motivo = (
                f"Segunda pasada: asignadas {assigned} de {target} con el stock "
                "remanente después de atender las solicitudes completas"
            )

        report_row.update(
            {
                "CANTIDAD_ASIGNADA": assigned,
                "CANTIDAD_FALTANTE": max(target - assigned, 0),
                "M3_ASIGNADO": assigned_m3,
                "M3_CAPACIDAD_ANTES": capacity_before,
                "M3_CAPACIDAD_DESPUES": capacity_after,
                "EXCEDE_CAPACIDAD_EN_ESTA_LINEA": False,
                "PASA_CAPACIDAD": not capacity_limited,
                "TAREAS_ANTES": task_before_second_pass,
                "TAREAS_GENERADAS": len(second_allocations),
                "TAREAS_ACUMULADAS": tasks_used,
                "PASA_TAREAS": not task_limited,
                "ORIGENES_USADOS": " | ".join(
                    f"{source}:{quantity}"
                    for source, quantity in second_allocations
                ),
                "VALUE": allocation_value_summary(
                    catalogs,
                    second_allocations,
                    sku,
                ),
                "STORAGE": allocation_storage_summary(
                    catalogs,
                    second_allocations,
                    sku,
                ),
                "TIPO_DE_CORTE": tipo_corte,
                "DETALLE_MOTIVO": detalle_motivo,
            }
        )
        for source in config.origin_warehouses:
            report_row.update(
                {
                    f"STOCK_ANTES_{source}": second_pass_before[source],
                    f"BLOQUEO_REGIONAL_{source}": second_pass_blocks[source],
                    f"VALUE_{source}": source_value_category(catalogs, source, sku),
                    f"STORAGE_{source}": source_storage_type(catalogs, source, sku),
                    f"ASIGNADO_{source}": assigned_by_origin[source],
                    f"STOCK_REMANENTE_{source}": stock_remaining[(source, sku)],
                }
            )

    capacity_rows: list[dict[str, Any]] = []
    destinations = sorted({row["WAREHOUSE_DESTINATION"] for row in prepared})
    for destination in destinations:
        store = catalogs.stores.get(destination, {})
        capacity_rows.append(
            {
                "WAREHOUSE_DESTINATION": destination,
                "WAREHOUSE_NAME": store.get("warehouse_name", ""),
                "CITY": store.get("city", ""),
                "CAPACIDAD_M3": catalogs.store_capacity.get(
                    destination, config.default_store_capacity_m3
                ),
                "M3_CONTABILIZADO_CAPACIDAD": cap_used_normal[destination],
                "M3_TOTAL_ASIGNADO_INCLUYE_GOLDEN": actual_m3_by_store[destination],
                "CAPACIDAD_CERRADA": cap_closed[destination],
                "CAPACIDAD_SUPERADA_POR_LINEA": cap_crossed_by_line[destination],
            }
        )

    return PlanningResult(
        base_rows=base_rows,
        allocation_rows=allocation_rows,
        capacity_rows=capacity_rows,
        warnings=catalogs.warnings,
        tasks_used=tasks_used,
        max_tasks=config.max_tasks,
    )


# %% Outputs

def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def excel_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and not math.isfinite(value):
        return ""
    return value


def write_rectangular_sheet(
    workbook,
    worksheet,
    headers: list[str],
    rows: Iterable[dict[str, Any]],
    header_format,
    boolean_format,
) -> int:
    worksheet.write_row(0, 0, headers, header_format)
    row_count = 0
    max_width = [len(header) for header in headers]
    for row_number, record in enumerate(rows, start=1):
        row_count += 1
        values = [excel_safe(record.get(header, "")) for header in headers]
        for column, value in enumerate(values):
            if isinstance(value, bool):
                worksheet.write_boolean(row_number, column, value, boolean_format)
            else:
                worksheet.write(row_number, column, value)
            max_width[column] = min(max(max_width[column], len(str(value))), 42)
    if row_count:
        worksheet.autofilter(0, 0, row_count, len(headers) - 1)
    worksheet.freeze_panes(1, 0)
    for column, width in enumerate(max_width):
        worksheet.set_column(column, column, max(width + 2, 10))
    return row_count


def write_planning_report(
    path: Path,
    result: PlanningResult,
    config: Config,
    run_date: date,
    input_filename: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(str(path), {"constant_memory": True})
    try:
        title_format = workbook.add_format(
            {
                "bold": True,
                "font_size": 16,
                "font_color": "#FFFFFF",
                "bg_color": "#222222",
                "align": "left",
                "valign": "vcenter",
            }
        )
        section_format = workbook.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#FF5A5F"}
        )
        header_format = workbook.add_format(
            {
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": "#333333",
                "border": 1,
                "text_wrap": True,
                "valign": "top",
            }
        )
        boolean_format = workbook.add_format({"align": "center"})
        integer_format = workbook.add_format({"num_format": "#,##0"})
        decimal_format = workbook.add_format({"num_format": "0.000000"})

        status_counts = Counter(row["TIPO_DE_CORTE"] for row in result.base_rows)
        units_by_source = Counter()
        tasks_by_source = Counter()
        for row in result.allocation_rows:
            source = row["WAREHOUSE_SOURCE"]
            units_by_source[source] += row["QUANTITY"]
            tasks_by_source[source] += 1

        target_units = sum(row["CANTIDAD_OBJETIVO"] for row in result.base_rows)
        assigned_units = sum(row["CANTIDAD_ASIGNADA"] for row in result.base_rows)
        total_input_rows = sum(
            row.get("CANTIDAD_FILAS_INPUT", 1) for row in result.base_rows
        )
        duplicate_keys = sum(
            row.get("CANTIDAD_FILAS_INPUT", 1) > 1 for row in result.base_rows
        )
        conflicting_duplicate_keys = sum(
            bool(row.get("DUPLICADO_CONFLICTIVO", False)) for row in result.base_rows
        )
        metrics = [
            ("FECHA_EJECUCION", run_date.strftime("%d-%m-%Y")),
            ("ARCHIVO_INPUT", input_filename),
            ("ORIGENES_CONFIGURADOS", ", ".join(map(str, config.origin_warehouses))),
            ("FILAS_CSV_INPUT", total_input_rows),
            ("REQUERIMIENTOS_UNICOS", len(result.base_rows)),
            ("FILAS_DUPLICADAS_CONSOLIDADAS", total_input_rows - len(result.base_rows)),
            ("LLAVES_DESTINO_SKU_DUPLICADAS", duplicate_keys),
            ("LLAVES_DUPLICADAS_CONFLICTIVAS", conflicting_duplicate_keys),
            (
                "REQUERIMIENTOS_CON_DEMANDA",
                sum(row["CANTIDAD_OBJETIVO"] > 0 for row in result.base_rows),
            ),
            ("UNIDADES_OBJETIVO", target_units),
            ("UNIDADES_ASIGNADAS", assigned_units),
            ("UNIDADES_FALTANTES", max(target_units - assigned_units, 0)),
            ("TAREAS_GENERADAS", result.tasks_used),
            ("MAX_TASKS", result.max_tasks),
            ("ADVERTENCIAS_DATOS", len(result.warnings)),
        ]

        summary = workbook.add_worksheet("RESUMEN")
        summary.hide_gridlines(2)
        summary.set_row(0, 28)
        summary.merge_range("A1:I1", "REPORTE DE PLANEACIÓN DE TRANSFERENCIAS", title_format)
        summary.write("A3", "MÉTRICAS", section_format)
        summary.write("D3", "TIPO DE CORTE", section_format)
        summary.write("G3", "ASIGNACIÓN POR SOURCE", section_format)
        for index, (metric, value) in enumerate(metrics, start=3):
            summary.write(index, 0, metric)
            summary.write(index, 1, value, integer_format if isinstance(value, int) else None)
        for index, (status, count) in enumerate(sorted(status_counts.items()), start=3):
            summary.write(index, 3, status)
            summary.write(index, 4, count, integer_format)
        for index, source in enumerate(config.origin_warehouses, start=3):
            summary.write(index, 6, source)
            summary.write(index, 7, tasks_by_source[source], integer_format)
            summary.write(index, 8, units_by_source[source], integer_format)
        summary.write(2, 7, "TAREAS", section_format)
        summary.write(2, 8, "UNIDADES", section_format)
        summary.set_column("A:A", 32)
        summary.set_column("B:B", 20)
        summary.set_column("D:D", 48)
        summary.set_column("E:E", 14)
        summary.set_column("G:I", 18)

        capacity_start = max(len(metrics), len(status_counts), len(config.origin_warehouses)) + 6
        summary.write(capacity_start, 0, "CAPACIDAD POR TIENDA", section_format)
        capacity_headers = list(result.capacity_rows[0]) if result.capacity_rows else []
        if capacity_headers:
            for col, header in enumerate(capacity_headers):
                summary.write(capacity_start + 1, col, header, header_format)
            for r_index, record in enumerate(result.capacity_rows, start=capacity_start + 2):
                for col, header in enumerate(capacity_headers):
                    value = record.get(header, "")
                    fmt = decimal_format if "M3" in header else None
                    summary.write(r_index, col, excel_safe(value), fmt)

        warning_start = capacity_start + len(result.capacity_rows) + 4
        summary.write(warning_start, 0, "ADVERTENCIAS DE CALIDAD", section_format)
        if result.warnings:
            for idx, warning in enumerate(result.warnings, start=warning_start + 1):
                summary.write(idx, 0, warning)
        else:
            summary.write(warning_start + 1, 0, "Sin advertencias")

        base = workbook.add_worksheet("BASE_TRANSFERS")
        base.hide_gridlines(2)
        base_headers = list(result.base_rows[0]) if result.base_rows else []
        if base_headers:
            base_count = write_rectangular_sheet(
                workbook,
                base,
                base_headers,
                result.base_rows,
                header_format,
                boolean_format,
            )
            status_col = base_headers.index("TIPO_DE_CORTE")
            base.conditional_format(
                1,
                status_col,
                base_count,
                status_col,
                {
                    "type": "text",
                    "criteria": "begins with",
                    "value": "OK",
                    "format": workbook.add_format(
                        {"bg_color": "#C6EFCE", "font_color": "#006100"}
                    ),
                },
            )
            base.conditional_format(
                1,
                status_col,
                base_count,
                status_col,
                {
                    "type": "text",
                    "criteria": "containing",
                    "value": "CORTE",
                    "format": workbook.add_format(
                        {"bg_color": "#FFC7CE", "font_color": "#9C0006"}
                    ),
                },
            )

        detail = workbook.add_worksheet("DETALLE_ASIGNACION")
        detail.hide_gridlines(2)
        write_rectangular_sheet(
            workbook,
            detail,
            OUTPUT_COLUMNS,
            result.allocation_rows,
            header_format,
            boolean_format,
        )
    finally:
        workbook.close()


def create_output_files(
    result: PlanningResult,
    config: Config,
    run_date: date,
    input_filename: str,
    output_dir: Path,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"Reporte_Planeacion_{run_date:%d-%m-%Y}.xlsx"
    write_planning_report(report_path, result, config, run_date, input_filename)
    paths = [report_path]

    rows_by_source: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in result.allocation_rows:
        rows_by_source[row["WAREHOUSE_SOURCE"]].append(row)

    for source in config.origin_warehouses:
        rows = rows_by_source.get(source, [])
        if not rows and not config.generate_empty_source_files:
            continue
        path = output_dir / f"BulkCD_{source}.csv"
        write_csv(path, rows, OUTPUT_COLUMNS)
        paths.append(path)
    return paths


# %% Orquestación

def run_pipeline(
    config: Config,
    *,
    local_data_transfers_path: str | Path | None = None,
    local_plan_path: str | Path | None = None,
    upload_to_drive: bool = True,
) -> dict[str, Any]:
    if not config.origin_warehouses:
        raise ValueError("origin_warehouses no puede estar vacío")
    if len(set(config.origin_warehouses)) != len(config.origin_warehouses):
        raise ValueError("origin_warehouses contiene duplicados")
    if config.max_tasks < 0:
        raise ValueError("max_tasks no puede ser negativo")

    run_date = parse_run_date(config)
    date_text = run_date.strftime("%d-%m-%Y")
    work_dir = Path(config.local_work_dir)
    input_dir = work_dir / "inputs" / date_text
    output_dir = work_dir / "outputs" / date_text
    input_dir.mkdir(parents=True, exist_ok=True)
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    drive = None
    plan_drive_metadata: dict[str, Any] | None = None
    if local_data_transfers_path and local_plan_path:
        data_path = Path(local_data_transfers_path)
        plan_path = Path(local_plan_path)
        if not data_path.exists() or not plan_path.exists():
            raise FileNotFoundError("No existen los archivos locales de prueba")
    else:
        if not upload_to_drive:
            raise ValueError(
                "Sin archivos locales, upload_to_drive debe ser True para descargar de Drive"
            )
        drive = authenticate_drive()
        spreadsheet_id = extract_drive_id(
            config.data_transfers_spreadsheet, "data_transfers_spreadsheet"
        )
        plans_folder_id = extract_drive_id(config.tr_plans_folder, "tr_plans_folder")
        data_path = input_dir / "DATA_TRANSFERS.xlsx"
        download_drive_file(drive, spreadsheet_id, data_path)
        plan_drive_metadata = find_daily_plan_file(drive, plans_folder_id, run_date)
        plan_path = input_dir / plan_drive_metadata["name"]
        download_drive_file(drive, plan_drive_metadata["id"], plan_path)

    print(f"Fecha de ejecución: {date_text}")
    print(f"DATA_TRANSFERS: {data_path.name}")
    print(f"Plan diario: {plan_path.name}")
    print(f"Orígenes: {list(config.origin_warehouses)}")

    catalogs = load_catalogs(data_path, config)
    plan_read = read_plan_csv(plan_path, config)
    result = plan_transfers(plan_read.rows, catalogs, config)
    result.warnings.extend(plan_read.warnings)
    local_files = create_output_files(
        result, config, run_date, plan_path.name, output_dir
    )

    uploaded: list[dict[str, Any]] = []
    remote_folder: dict[str, Any] | None = None
    if upload_to_drive:
        if drive is None:
            drive = authenticate_drive()
        plans_folder_id = extract_drive_id(config.tr_plans_folder, "tr_plans_folder")
        output_root = find_or_create_folder(
            drive, plans_folder_id, config.remote_output_root_name
        )
        remote_folder = find_or_create_folder(drive, output_root["id"], date_text)
        for path in local_files:
            mime = (
                "text/csv"
                if path.suffix.lower() == ".csv"
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            uploaded.append(
                upload_or_replace_file(
                    drive,
                    path,
                    remote_folder["id"],
                    mime,
                    config.replace_same_day_outputs,
                )
            )

    status_counts = Counter(row["TIPO_DE_CORTE"] for row in result.base_rows)
    print("\nProceso finalizado")
    print(f"Filas leídas del CSV: {plan_read.input_row_count:,}")
    print(f"Requerimientos únicos: {len(result.base_rows):,}")
    print(f"Filas duplicadas consolidadas: {plan_read.duplicate_row_count:,}")
    print(f"Llaves duplicadas: {plan_read.duplicate_key_count:,}")
    print(
        "Llaves duplicadas con métricas diferentes: "
        f"{plan_read.conflicting_duplicate_key_count:,}"
    )
    print(f"Asignaciones/tareas: {result.tasks_used:,} de {result.max_tasks:,}")
    for source in config.origin_warehouses:
        source_rows = sum(
            row["WAREHOUSE_SOURCE"] == source for row in result.allocation_rows
        )
        source_units = sum(
            row["QUANTITY"]
            for row in result.allocation_rows
            if row["WAREHOUSE_SOURCE"] == source
        )
        print(f"Source {source}: {source_rows:,} líneas / {source_units:,} unidades")
    print("Tipos de corte:")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count:,}")
    if result.warnings:
        print(f"Advertencias de calidad: {len(result.warnings)} (ver RESUMEN)")
    print("Archivos locales:")
    for path in local_files:
        print(f"  {path}")
    if uploaded:
        print("Archivos guardados en Drive:")
        for item in uploaded:
            print(f"  {item.get('name')}: {item.get('webViewLink', item.get('id'))}")

    return {
        "run_date": run_date,
        "result": result,
        "local_files": local_files,
        "uploaded_files": uploaded,
        "remote_folder": remote_folder,
        "plan_drive_metadata": plan_drive_metadata,
    }


def main() -> None:
    run_pipeline(CONFIG)


if __name__ == "__main__":
    main()
